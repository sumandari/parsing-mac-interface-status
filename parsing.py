import argparse
import ast
import re
import sys
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook, Workbook
from mac_vendor_lookup import AsyncMacLookup, MacLookup


FIELDNAMES = [
    'hostname',
    'port',
    'name',
    'status',
    'vlan',
    'duplex',
    'speed',
    'macAddress',
    'macVendor'
]
STATUS = ['connected', 'up']


def check_if_file_exist(filename):
    """The name already describes the function."""

    file = Path(filename)
    if not file.is_file():
        print(f"File {filename} does not exist.")
        return None
    return file


def check_if_dir_exist(dirname):
    """The name already describes the function."""

    dir = Path(dirname)
    if not dir.is_dir():
        print(f"Directory {dir} does not exist.")
        sys.exit(1)


def get_port_alias(port):
    """Get the alias name for a port.

    e.g port: Ethernet123, mac_alias: Et123
    """

    port_name_pattern = re.findall(
        r'([a-zA-Z]+)([0-9/]+)',
        port
    )
    return port_name_pattern[0][0][0:2] + port_name_pattern[0][1]


def init_argparse() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        # usage="parsing.py [OPTION] [fileiface1,filemac1]...",
        description="Parse mac address and interface data."
    )
    parser.add_argument(
        'files', metavar='files', nargs='+',
        help='files to parse. eg. file_iface1,file_mac1 '
    )
    parser.add_argument(
        '-d', '--directory', type=str,
        default=str(Path().absolute()),
        help='assign local directory for output excel file.'
    )
    parser.add_argument(
        '-f', '--filter', type=str,
        help='filter mac addresses. Generates the excel output with '
        'only the specified mac address listed in a txt file.'
    )
    return parser


def lookup_mac_vendor(macaddress):
    """Get the vendor name."""

    mac = MacLookup()
    return mac.lookup(macaddress)


def parsing_filter(file):
    print("Parsing list of mac addresses filter...")

    filters = []
    with open(file) as f:
        lines = f.readlines()
        filters = list(map(str.strip, lines))
    return filters


def parsing_mac_address(file):

    print("Start parsing mac address...")

    macs = {}
    with open(file) as f:
        lines = f.readlines()
        for _ in lines[5:-1]:
            if _.startswith("Total Mac Addresses for this criterion:"):
                break
            line = _.split()
            vlan = line[0]
            mac = line[1]
            mac_type = line[2]
            port = line[3]
            # There are "STATIC" and "DYNAMIC" mac addresses.
            if vlan and mac and port:
                try:
                    mac_vendor = lookup_mac_vendor(mac)
                except Exception as e:
                    print('exception', e)
                    mac_vendor = ''

                if mac_type == 'DYNAMIC':
                    if port in macs:
                        mac_vendor = (
                            f"{macs[port]['macVendor']}; {mac_vendor}"
                        )
                        mac = f"{macs[port]['mac']}; {mac}"
                        vlan = f"{macs[port]['vlan']}; {vlan}"

                macs[port] = {
                    'vlan': vlan,
                    'mac': mac,
                    'macVendor': mac_vendor
                }
    return macs


def parsing_interface(iface, mac_address):
    """Parsing files"""

    print("Start parsing interface...")

    result = []
    with open(iface) as f:
        line = f.readline()
        data = ast.literal_eval(line.strip()[1:-1])
        if not isinstance(data, dict):
            print(f"Invalid file format: {iface}.")
            exit(1)

        ifaces = data.get('ansible_facts', None)
        hostname = ifaces.get('ansible_net_hostname', None)

        try:
            interfaces = ifaces['ansible_network_resources']['interfaces']
            iface_data = ifaces['ansible_net_interfaces']
        except KeyError as e:
            print("Invalid interfaces data.", e)
            exit(1)

        parsed_macs = parsing_mac_address(mac_address)
        for interface in interfaces:
            try:
                port = interface['name']
                status = iface_data[port].get('operstatus', None)
                if not status or not status in STATUS:
                    continue

                name = iface_data[port].get('description', None)
                duplex = iface_data[port].get('duplex', None)
                speed = iface_data[port].get('bandwidth', None)

                vlan = None
                mac = None
                mac_vendor = None

                port_alias = get_port_alias(port)
                if port_alias in parsed_macs:
                    vlan = parsed_macs[port_alias]['vlan']
                    mac = parsed_macs[port_alias]['mac']
                    mac_vendor = parsed_macs[port_alias]['macVendor']

                result.append((
                    hostname,
                    port,
                    name,
                    status,
                    vlan,
                    duplex,
                    speed,
                    mac,
                    mac_vendor
                ))

            except KeyError as e:
                print(f"Invalid interfaces data in port {port}.", e.__str__())
    return result


def save_to_xlsx(iface, mac_addres, output, filter):
    """Save the parsed interfaces and mac addresses data in xlsx file."""

    data = parsing_interface(iface, mac_addres)
    hostname = f"{data[0][0]}"

    wb = load_workbook(output)
    # remove default empty sheet.
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    # create new sheet with hostname as sheetname.
    wb.create_sheet(title=hostname)
    sheet = wb[hostname]
    sheet.append(FIELDNAMES)

    if filter:
        filters = parsing_filter(filter)
        for row in data:
            (
                hostname,
                port,
                name,
                status,
                vlan,
                duplex,
                speed,
                mac,
                mac_vendor
            ) = row
            if mac in filters:
                sheet.append(row)
    else:
        for row in data:
            sheet.append(row)
    wb.save(output)
    print(f"Saved data {iface} and {mac_addres} in {output}")


def get_worksheet_name(dirname):
    utcnow = datetime.utcnow()
    minutes = utcnow.minute // 10 + 1
    filename = (
        f"report_{utcnow.year}-{utcnow.month}"
        f"-{utcnow.day}_{utcnow.hour}-{minutes}"
    )
    suffix = ".xlsx"
    return str(Path(dirname, filename).with_suffix(suffix))


if __name__ == "__main__":
    args_parser = init_argparse()
    args = args_parser.parse_args()

    check_if_dir_exist(args.directory)
    output = get_worksheet_name(args.directory)
    if not check_if_file_exist(output):
        wb = Workbook()
        wb.save(output)
        print(f"Created file {output} to save the data.")

    filter = args.filter
    if filter:
        filter = check_if_file_exist(filter)
    for f in args.files:
        files = f.split(',')
        iface = check_if_file_exist(files[0])
        mac_address = check_if_file_exist(files[1])

        if iface and mac_address:
            save_to_xlsx(iface, mac_address, output, filter)
